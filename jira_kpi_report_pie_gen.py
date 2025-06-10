from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from collections import defaultdict

def generate_sprint_report_with_percent_pies(filepath):
    wb = load_workbook(filepath)
    ws_summary = wb["Summary"]

    if "ChartData" in wb.sheetnames:
        del wb["ChartData"]
    ws_data = wb.create_sheet("ChartData")
    ws_data.sheet_state = "hidden"

    rows = list(ws_summary.iter_rows(values_only=True))
    team_blocks = [(row[0].strip(), i) for i, row in enumerate(rows) if isinstance(row[0], str) and row[0].strip().endswith("TEAM")]
    team_bounds = [(name, start, team_blocks[i + 1][1] if i + 1 < len(team_blocks) else len(rows)) for i, (name, start) in enumerate(team_blocks)]

    chart_anchor = 5
    chart_spacing = 18
    row_cursor = 1

    for team_name, start, end in team_bounds:
        task_totals = defaultdict(int)
        current_category = None
        status_col_index = None

        for r in range(start, min(end, start + 5)):
            if rows[r][0] == "Task Category / Status":
                try:
                    status_col_index = rows[r].index("Status Totals")
                except ValueError:
                    status_col_index = None
                break

        if status_col_index is None:
            continue

        for i in range(start, end):
            row = rows[i]
            if not row or all(cell is None for cell in row):
                continue
            first = row[0]
            if isinstance(first, str) and first.strip() not in (
                "To Do", "In Development", "Completed", "Declined", "Task Category / Status", "TOTAL"):
                current_category = first.strip()
            elif (
                current_category
                and isinstance(first, str)
                and first.strip() in ("To Do", "In Development", "Completed", "Declined")
            ):
                val = row[status_col_index]
                if isinstance(val, (int, float, str)) and str(val).isdigit():
                    task_totals[current_category] += int(val)

        filtered_totals = {k: v for k, v in task_totals.items() if v > 0}
        if not filtered_totals:
            continue

        start_row = row_cursor
        for cat, val in filtered_totals.items():
            ws_data.cell(row=row_cursor, column=1, value=cat)
            ws_data.cell(row=row_cursor, column=2, value=val)
            row_cursor += 1
        end_row = row_cursor - 1

        label_ref = Reference(ws_data, min_col=1, min_row=start_row, max_row=end_row)
        data_ref = Reference(ws_data, min_col=2, min_row=start_row, max_row=end_row)

        chart = PieChart()
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(label_ref)
        chart.title = f"{team_name} Task Contribution"
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showVal = False
        chart.dataLabels.showLeaderLines = False
        chart.dataLabels.showLegendKey = False

        ws_summary.add_chart(chart, f"Y{chart_anchor}")
        chart_anchor += chart_spacing

    wb.save(filepath)

if __name__ == "__main__":
    generate_sprint_report_with_percent_pies("sprint_report.xlsx")

if __name__ == '__main__':
    generate_sprint_report_with_percent_pies('sprint_report.xlsx')
