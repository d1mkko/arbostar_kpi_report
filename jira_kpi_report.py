#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from jira import JIRA
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import mock data generator for BA TEAM
try:
    from mock_ba_data import generate_mock_ba_data
except ImportError:
    # Define a fallback mock data generator in case the import fails
    def generate_mock_ba_data():
        print("WARNING: mock_ba_data.py not found, using empty mock data")
        return {
            'BA TEAM': {
                category: {
                    'prev': {member: {'To Do': 0, 'In Development': 0, 'Completed': 0, 'Declined': 0} 
                            for member in ['Bohdan Kucher', 'Polina Reminna', 'Stepan Zhukevych']},
                    'pre_prev': {member: {'To Do': 0, 'In Development': 0, 'Completed': 0, 'Declined': 0}
                                for member in ['Bohdan Kucher', 'Polina Reminna', 'Stepan Zhukevych']},
                    'tasks': {'prev': {}, 'pre_prev': {}}
                }
                for category in ['ASAP BA changes', 'Change requests BA']
            }
        }

# Import mock data generator for AMA TEAM
try:
    from mock_ama_data import generate_mock_ama_data
except ImportError:
    # Define a fallback mock data generator in case the import fails
    def generate_mock_ama_data():
        print("WARNING: mock_ama_data.py not found, using empty mock data")
        return {
            'AMA TEAM': {
                category: {
                    'prev': {member: {'To Do': 0, 'In Development': 0, 'Completed': 0, 'Cancelled': 0} 
                            for member in ['Andriy Momot', 'Arthur Hlushko', 'Denys Honchar', 
                                        'Iliya Sozonenko', 'Oleg Nekrasov', 'Oleksandr Korneiko', 
                                        'Oleksii Petrov']},
                    'pre_prev': {member: {'To Do': 0, 'In Development': 0, 'Completed': 0, 'Cancelled': 0}
                                for member in ['Andriy Momot', 'Arthur Hlushko', 'Denys Honchar', 
                                            'Iliya Sozonenko', 'Oleg Nekrasov', 'Oleksandr Korneiko', 
                                            'Oleksii Petrov']},
                    'tasks': {'prev': {}, 'pre_prev': {}}
                }
                for category in ['ASAP Changes', 'Change Requests', 'BugFixes']
            }
        }

# Import mock data generator for LDT, TWA, and CWT teams
try:
    from mock_ldt_twa_cwt_data import generate_mock_data
except ImportError:
    # Define a fallback mock data generator in case the import fails
    def generate_mock_data():
        print("WARNING: mock_ldt_twa_cwt_data.py not found, using empty mock data")
        return {
            'LDT TEAM': {category: {'prev': {}, 'pre_prev': {}, 'tasks': {'prev': {}, 'pre_prev': {}}}
                    for category in ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes', 'Client PDF', 'Migration']},
            'TWA TEAM': {category: {'prev': {}, 'pre_prev': {}, 'tasks': {'prev': {}, 'pre_prev': {}}}
                    for category in ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes', 'Client PDF', 'Migration']},
            'CWT TEAM': {category: {'prev': {}, 'pre_prev': {}, 'tasks': {'prev': {}, 'pre_prev': {}}}
                    for category in ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes', 'Client PDF', 'Migration']},
        }

# Jira connection details
JIRA_SERVER = 'https://arbostar.atlassian.net'
JIRA_EMAIL = '<your_email>'
JIRA_API_TOKEN = '<Your_API_Token>' # Kept from your uploaded file

# Team members
# LDT TEAM
LDT_TEAM_MEMBERS = ['Andrew Belousov', 'Ivan Stepaniuk', 'serg levch']
# TWA TEAM
TWA_TEAM_MEMBERS = ['Anton Rozonenko', 'Anton Shelekhvost', 'Anton Shynkarenko', 
                    'Dmytro Yurchenko', 'Ivan Borovets', 'Maksim Levchenko', 
                    'Michael Parandiy', 'Oleg Lats', 'Oleg Nekrasov', 
                    'Oleksii Petrov', 'Roman Dubovka', 'Zubkov Pavlo']
# CWT TEAM
CWT_TEAM_MEMBERS = ['Sergey Chernov']
# BA TEAM
BA_TEAM_MEMBERS = ['Bohdan Kucher', 'Polina Reminna', 'Stepan Zhukevych']
# AMA TEAM
AMA_TEAM_MEMBERS = ['Andriy Momot', 'Arthur Hlushko', 'Denys Honchar', 
                'Iliya Sozonenko', 'Oleg Nekrasov', 'Oleksandr Korneiko', 
                'Oleksii Petrov']

# All teams and their members
TEAMS = {
    'LDT TEAM': LDT_TEAM_MEMBERS,
    'TWA TEAM': TWA_TEAM_MEMBERS,
    'CWT TEAM': CWT_TEAM_MEMBERS,
    'BA TEAM': BA_TEAM_MEMBERS,
    'AMA TEAM': AMA_TEAM_MEMBERS
}

# Date ranges
# IMPORTANT: These must consistently match the JQL's relative date format (e.g., "-21d")
PREV_SPRINT_START = "-21d" 
PREV_SPRINT_END = "-7d"
PRE_PREV_SPRINT_START = "-42d" 
PRE_PREV_SPRINT_END = "-28d"   

# Template path
OUTPUT_PATH = "sprint_report.xlsx"

# Status mappings for LDT, TWA, and CWT teams
STATUS_MAPPING = {
    'TO_DO': ['To Do'],
    'IN_DEV': ['⚙️ In Progress', 'Paused', 'Reopen', '🎯 Resolved', 'DEV', 'Merge to Staging', 
            'Staging', 'HOTFIX', 'Merge to Master'],
    'COMPLETED': ['MASTER', 'Ready for release']
}

# BA team status categories
BA_STATUS_MAPPING = {
    'TO_DO': ['Backlog', 'New change', 'Need details', 'Awaiting approval', 'Investigate', 'Ready to do', 'Ready to Do'],
    'IN_DEV': ['In BA progress', 'In BA Progress', 'Waiting for feedback', 'Waiting for Feedback', 'Ready for development', 'In development', 'In Development',
            'Scheduled for release', 'In support'],
    'COMPLETED': ['Closed', 'Updating in knowledge base'],
    'DECLINED': ['Declined']  # Renamed from CANCELLED to DECLINED for consistency
}

# AMA team status categories
AMA_STATUS_MAPPING = {
    'TO_DO': ['To Do'],
    'IN_DEV': ['In Progress', 'Pending', 'Code Review', 'Ready to test', 'Test passed', 'Test pre-release', 
            'Merge to Master', 'Staging', 'Paused', '⚙️ In Progress', '🎯 Resolved', 'Reopen'],
    'COMPLETED': ['Ready for release', 'Released', 'MASTER'],
    'CANCELLED': ['Cancelled']  # New status category for AMA team
}

# Team-specific status mappings
TEAM_STATUS_MAPPINGS = {
    'BA TEAM': BA_STATUS_MAPPING,
    'AMA TEAM': AMA_STATUS_MAPPING,
    # Default mapping for other teams
    'LDT TEAM': STATUS_MAPPING,
    'TWA TEAM': STATUS_MAPPING,
    'CWT TEAM': STATUS_MAPPING
}

# Task categories - RESTORED TO ORIGINAL JQL QUERIES FROM YOUR FILE
TASK_CATEGORIES = {
    'ASAP Changes': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND issuetype="Change request" AND Release[Dropdown]=ASAP  and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d"  and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}',
        'ama_query': '(project = "AMA") AND issuetype="Change request" and Release[Dropdown]=ASAP and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("Ready to test", "Test passed", "Test pre-release", "Ready for release", "Released", "Cancelled") and assignee = {assignee}'
    },
    'Change Requests': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND issuetype="Change request" AND Release[Dropdown] IS EMPTY  and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d"  and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}',
        'ama_query': '(project = "AMA") AND issuetype="Change request" AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("Ready to test", "Test passed", "Test pre-release", "Ready for release", "Released", "Cancelled") and assignee = {assignee}'
    },
    'Tech. Tasks': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND issuetype = \'Task\' AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d"  and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}',
        'ama_query': '(project = "AMA") AND issuetype = "Task" AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("Ready to test", "Test passed", "Test pre-release", "Ready for release", "Released", "Cancelled") and assignee = {assignee}'
    },
    'BugFixes': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND issuetype = "Bug" AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d"  and statusCategoryChangedDate <= "-7d" and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}',
        'ama_query': '(project = "AMA") AND issuetype = "Bug" AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("Ready to test", "Test passed", "Test pre-release", "Ready for release", "Released", "Cancelled") and assignee = {assignee}'
    },
    'Client PDF': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND "Epic Link" = TWA-3303 AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}'
    },
    'Migration': {
        'query': '(project = "TWA" OR project = "LDT" OR project = "CWT") AND "Epic Link" = TWA-3306 AND Release[Dropdown] IS EMPTY and statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" and  status not in ("DEV", "Merge to Staging", "Staging", "HOTFIX", "Merge to Master", "MASTER", "Ready for release") and assignee = {assignee}'
    },
    # For BA team categories - now just one category
    'Change Requests BA': {
        'query': '(project = "Features and Ideas")  AND statusCategoryChangedDate >= "-21d" and statusCategoryChangedDate <= "-7d" AND assignee = {assignee}'
    }
}

# Define teams that use task categories
TEAM_CATEGORIES = {
    'LDT TEAM': ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes', 'Client PDF'],
    'TWA TEAM': ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes', 'Client PDF', 'Migration'],
    'CWT TEAM': ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes'],
    'BA TEAM': ['Change Requests BA'],
    'AMA TEAM': ['ASAP Changes', 'Change Requests', 'Tech. Tasks', 'BugFixes']
}

# Flag to use mock data for teams
USE_MOCK_BA_DATA = False
USE_MOCK_AMA_DATA = False
USE_MOCK_OTHER_DATA = False  # Flag to use mock data for LDT, TWA, and CWT teams

def connect_to_jira():
    """Connect to Jira using API token"""
    print("Connecting to Jira...")
    try:
        jira = JIRA(server=JIRA_SERVER, basic_auth=(JIRA_EMAIL, JIRA_API_TOKEN))
        print("Connected successfully!")
        return jira
    except Exception as e:
        print(f"Failed to connect to Jira: {e}")
        sys.exit(1)

def create_jql_query(category, date_start, date_end, assignee=None, team_name=None):
    """
    Create JQL query based on task category and date range.
    date_start and date_end are expected to be *relative date strings* (e.g., "-21d")
    from the global variables.
    """
    category_info = TASK_CATEGORIES[category]
    
    # If the category has a predefined query, use it
    if 'query' in category_info:
        # Use special AMA query if team is AMA and the category has an AMA-specific query
        if team_name == 'AMA TEAM' and 'ama_query' in category_info:
            query = category_info['ama_query']
        else:
            query = category_info['query']
            
        # Replace the placeholder with the actual assignee.
        # The date parts ({date_start}, {date_end}) are NOT in these JQLs, as per user's original query.
        # The relative dates are hardcoded in the TASK_CATEGORIES JQL themselves.
        if assignee:
            # Format the assignee name with proper escaping for JQL
            formatted_assignee = f'"{assignee}"'
            query = query.replace('{assignee}', formatted_assignee)
        return query
    
    # This legacy query building code should ideally not be reached with current TASK_CATEGORIES structure
    boards = category_info.get('boards', [])
    if team_name == 'BA TEAM':
        boards = category_info.get('boards', [])
    
    boards_clause = ' OR '.join([f'project = "{board}"' for board in boards])
    boards_clause = f'({boards_clause})'
    
    query = f"{boards_clause}"
    
    if 'type' in category_info:
        query += f" AND issuetype = '{category_info['type']}'"
    
    if 'release' in category_info:
        if 'release_not' in category_info:
            query += f" AND (cf[10102] != '{category_info['release_not']}' OR cf[10102] is EMPTY)"
        else:
            query += f" AND cf[10102] = '{category_info['release']}'"
    
    if 'epic' in category_info:
        query += f" AND \"Epic Link\" = {category_info['epic']}"
    
    query += f" AND updated >= '{date_start}' AND updated <= '{date_end}'"
    
    if assignee:
        formatted_name = f'"{assignee}"'
        query += f" AND assignee = {formatted_name}"
    
    return query

def get_tasks_for_period(jira, category, date_start_relative, date_end_relative, assignee, team_name):
    """
    Get tasks for a specific period, category, and team member,
    including story points from customfield_10149.
    date_start_relative and date_end_relative are relative date strings (e.g., "-21d").
    """
    all_tasks = []
    
    # Skip if this category doesn't apply to this team
    if category not in TEAM_CATEGORIES.get(team_name, []):
        return all_tasks
    
    # Pass the relative dates to create_jql_query, as the JQL itself uses relative dates.
    jql = create_jql_query(category, date_start_relative, date_end_relative, assignee, team_name)
    try:
        print(f"Executing JQL for {team_name}, {assignee}: {jql}")
        # Request customfield_10149 (Story Points)
        issues = jira.search_issues(jql, maxResults=500, fields='summary,status,assignee,customfield_10149')
        print(f"Found {len(issues)} issues for {assignee} in {team_name}")
        
        for issue in issues:
            status = issue.fields.status.name
            story_points = getattr(issue.fields, 'customfield_10149', 0.0)
            if story_points is None:
                story_points = 0.0
            
            # Get the status mapping for this team
            status_mapping = TEAM_STATUS_MAPPINGS.get(team_name, STATUS_MAPPING)
            
            if status in status_mapping['TO_DO']:
                status_category = "To Do"
            elif status in status_mapping['IN_DEV']:
                status_category = "In Development"
            elif status in status_mapping['COMPLETED']:
                status_category = "Completed"
            elif 'DECLINED' in status_mapping and status in status_mapping['DECLINED']:
                status_category = "Declined"
            elif 'CANCELLED' in status_mapping and status in status_mapping['CANCELLED']:
                status_category = "Cancelled"
            else:
                status_category = "Other"
                print(f"Warning: Status '{status}' for issue {issue.key} was not mapped to any category for team {team_name}")
            
            all_tasks.append({
                'Key': issue.key,
                'Summary': issue.fields.summary,
                'Status': status,
                'StatusCategory': status_category,
                'Assignee': assignee,
                'StoryPoints': story_points,
            })
    except Exception as e:
        print(f"Error in JQL query '{jql}': {e}")
        print(f"JQL: {jql}")
    
    return all_tasks

def get_tracked_time_for_period(jira, date_start_relative, date_end_relative, team_members):
    """
    Fetches all worklogs within a given period and aggregates time spent by each team member.
    Uses relative dates for the JQL query to fetch issues, and then filters in Python for robustness.
    Includes detailed logging.
    """
    tracked_time_by_member = {member: 0.0 for member in team_members}

    # Derive absolute dates from relative date strings for Python-side filtering
    current_system_time = datetime.now()
    def parse_relative_date_to_absolute(relative_str, base_date):
        """Helper to parse relative date strings like '-7d' to absolute dates."""
        if relative_str.startswith('-') and relative_str.endswith('d'):
            days_offset = int(relative_str[1:-1])
            return base_date - timedelta(days=days_offset) # Subtract for past dates
        # Fallback for absolute dates if format is YYYY-MM-DD (though we expect relative here)
        try:
            return datetime.strptime(relative_str, "%Y-%m-%d")
        except ValueError:
            print(f"WARNING: Unexpected date format '{relative_str}'. Cannot parse to absolute date for internal filtering.")
            return base_date # Return base date as a fallback if parsing fails
    
    start_date_obj_abs = parse_relative_date_to_absolute(date_start_relative, current_system_time).date()
    end_date_obj_abs = parse_relative_date_to_absolute(date_end_relative, current_system_time).date()
    
    print(f"\n--- Fetching Worklogs for Tracked Time ---")
    print(f"  System Time Used for Calculation: {current_system_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Report Period (relative for JQL): {date_start_relative} to {date_end_relative}")
    print(f"  Report Period (absolute for internal Python checks): {start_date_obj_abs} to {end_date_obj_abs}")

    all_projects = set()
    for team_name_key in TEAMS.keys():
        for category in TEAM_CATEGORIES.get(team_name_key, []):
            category_info = TASK_CATEGORIES.get(category)
            if category_info:
                import re
                # Extract project names from both 'query' and 'ama_query' using regex
                project_matches_query = re.findall(r'project\s*=\s*"([^"]+)"', category_info.get('query', ''))
                project_matches_ama_query = re.findall(r'project\s*=\s*"([^"]+)"', category_info.get('ama_query', ''))
                all_projects.update(project_matches_query)
                all_projects.update(project_matches_ama_query)
                
    project_jql_clause = ""
    if all_projects:
        project_jql_clause_content = " OR ".join([f'project = "{p}"' for p in all_projects])
        project_jql_clause = f"({project_jql_clause_content}) AND "
    else:
        print("WARNING: No projects found in TASK_CATEGORIES to limit worklog search. Searching all projects (this might be slow and permission-heavy).")
        project_jql_clause = "project is not EMPTY AND " # Fallback if no specific projects extracted


    assignee_list_for_jql = ', '.join([f'"{member}"' for member in team_members])
    
    # Construct JQL using the provided relative dates for worklogDate
    jql_broad_issues = (
        f"{project_jql_clause} worklogDate >= '{date_start_relative}' AND worklogDate <= '{date_end_relative}' "
        f"AND worklogAuthor in ({assignee_list_for_jql})"
    )
    
    print(f"Executing Broad JQL to find potential issues for worklogs: {jql_broad_issues}")
    
    try:
        # Request the 'worklog' field to get worklog details
        issues_to_check = jira.search_issues(jql_broad_issues, maxResults=False, fields='summary,worklog,assignee')
        print(f"Found {len(issues_to_check)} issues that might contain relevant worklogs by broad query.")

        if not issues_to_check:
            print("  No issues found by the broad query for worklogs. This might indicate a fundamental permission issue or no relevant activity in the period for these assignees.")

        for issue in issues_to_check:
            print(f"  Processing issue: {issue.key} - {issue.fields.summary}") 
            if hasattr(issue.fields, 'worklog') and hasattr(issue.fields.worklog, 'worklogs') and issue.fields.worklog.worklogs:
                print(f"    Issue {issue.key} has {len(issue.fields.worklog.worklogs)} worklog entries.")
                for worklog in issue.fields.worklog.worklogs:
                    worklog_author_display_name = worklog.author.displayName
                    worklog_started_str = worklog.started 
                    
                    try:
                        # Extract just the date part from the 'started' timestamp (e.g., '2023-05-10T12:00:00.000+0000' -> '2023-05-10')
                        worklog_date_obj = datetime.strptime(worklog_started_str.split('T')[0], "%Y-%m-%d").date()
                    except ValueError:
                        print(f"        WARNING: Could not parse worklog date '{worklog_started_str}' for issue {issue.key}. Skipping this worklog.")
                        continue 
                    
                    # Apply Python-side filtering to ensure worklogs fall strictly within the desired absolute date range
                    # AND ensure the author is one of the team members
                    if (start_date_obj_abs <= worklog_date_obj <= end_date_obj_abs) and \
                       (worklog_author_display_name in team_members):
                        
                        timespent_seconds = worklog.timeSpentSeconds
                        timespent_hours = timespent_seconds / 3600.0
                        tracked_time_by_member[worklog_author_display_name] += timespent_hours
                        print(f"        ✅ Worklog PROCESSED: Issue={issue.key}, Author='{worklog_author_display_name}', Date='{worklog_date_obj}', TimeSpentSeconds={timespent_seconds}, Added {timespent_hours:.2f} hours. Current Total for '{worklog_author_display_name}': {tracked_time_by_member[worklog_author_display_name]:.2f}")
                    else:
                        skip_reason = []
                        if not (start_date_obj_abs <= worklog_date_obj <= end_date_obj_abs):
                            skip_reason.append(f"date {worklog_date_obj} outside report period ({start_date_obj_abs} to {end_date_obj_abs})")
                        if worklog_author_display_name not in team_members:
                            skip_reason.append(f"author '{worklog_author_display_name}' not in team members list")
                        print(f"        ❌ Worklog SKIPPED: Issue={issue.key}, Author='{worklog_author_display_name}', Date='{worklog_date_obj}', TimeSpentSeconds={worklog.timeSpentSeconds}. Reason: {'; '.join(skip_reason)}")

    except Exception as e:
        print(f"ERROR: Failed to fetch worklogs with JQL '{jql_broad_issues}': {e}")
        import traceback
        traceback.print_exc() # Print full traceback for deeper debugging
    
    return tracked_time_by_member

def process_data(jira):
    """Process all data for categories and teams"""
    print("\n--- Entering process_data function ---") # Added print statement
    all_data = {}
    
    # Get BA TEAM mock data if enabled
    if USE_MOCK_BA_DATA:
        print("Using mock data for BA TEAM")
        ba_mock_data = generate_mock_ba_data()
        if ba_mock_data and 'BA TEAM' in ba_mock_data:
            all_data['BA TEAM'] = ba_mock_data['BA TEAM']
    
    # Get AMA TEAM mock data if enabled
    if USE_MOCK_AMA_DATA:
        print("Using mock data for AMA TEAM")
        ama_mock_data = generate_mock_ama_data()
        if ama_mock_data and 'AMA TEAM' in ama_mock_data:
            all_data['AMA TEAM'] = ama_mock_data['AMA TEAM']
    
    # Get mock data for LDT, TWA, and CWT teams if enabled
    if USE_MOCK_OTHER_DATA:
        print("Using mock data for LDT, TWA, and CWT teams")
        other_mock_data = generate_mock_data()
        for team_name in ['LDT TEAM', 'TWA TEAM', 'CWT TEAM']:
            if team_name in other_mock_data:
                all_data[team_name] = other_mock_data[team_name]
    
    # Process data for each team separately
    for team_name, team_members in TEAMS.items():
        # Skip teams if we're using mock data
        if ((team_name == 'BA TEAM' and USE_MOCK_BA_DATA) or 
            (team_name == 'AMA TEAM' and USE_MOCK_AMA_DATA) or 
            (team_name in ['LDT TEAM', 'TWA TEAM', 'CWT TEAM'] and USE_MOCK_OTHER_DATA)):
            print(f"Skipping live Jira data fetch for {team_name} due to mock data flag.")
            continue
            
        team_data = {}
        
        print(f"\n--- Calling get_tracked_time_for_period for {team_name} (Previous Sprint) ---")
        team_tracked_time_prev = get_tracked_time_for_period(jira, PREV_SPRINT_START, PREV_SPRINT_END, team_members)
        print(f"\n--- Calling get_tracked_time_for_period for {team_name} (Pre-Previous Sprint) ---")
        team_tracked_time_pre_prev = get_tracked_time_for_period(jira, PRE_PREV_SPRINT_START, PRE_PREV_SPRINT_END, team_members)

        # Store the aggregated tracked time directly at the team_data level
        team_data['aggregated_tracked_time'] = {
            'prev': team_tracked_time_prev,
            'pre_prev': team_tracked_time_pre_prev
        }

        for category in TEAM_CATEGORIES.get(team_name, list(TASK_CATEGORIES.keys())):
            print(f"Processing {category} for {team_name}...")
            category_data = {
                'prev': {},
                'pre_prev': {},
                'tasks': {'prev': {}, 'pre_prev': {}},
                'story_points': {'prev': {}, 'pre_prev': {}}, 
                # Removed 'tracked_time' from category_data as it's now aggregated at team_data level
            }
            
            for team_member in team_members:
                # Get previous sprint tasks (for counts and story points)
                # Use the hardcoded relative date strings for get_tasks_for_period
                prev_tasks = get_tasks_for_period(jira, category, PREV_SPRINT_START, PREV_SPRINT_END, team_member, team_name)
                
                # Get pre-previous sprint tasks (for counts and story points)
                # Use the hardcoded relative date strings for get_tasks_for_period
                pre_prev_tasks = get_tasks_for_period(jira, category, PRE_PREV_SPRINT_START, PRE_PREV_SPRINT_END, team_member, team_name)
                
                # Get the status categories for this team
                status_mapping = TEAM_STATUS_MAPPINGS.get(team_name, STATUS_MAPPING)
                status_categories = ['To Do', 'In Development', 'Completed']
                
                # Add special status categories if present in this team's mapping
                if 'DECLINED' in status_mapping:
                    status_categories.append('Declined')
                if 'CANCELLED' in status_mapping:
                    status_categories.append('Cancelled')
                
                # Count tasks by status for each period and team member
                prev_counts = {status: sum(1 for t in prev_tasks if t['StatusCategory'] == status) for status in status_categories}
                pre_prev_counts = {status: sum(1 for t in pre_prev_tasks if t['StatusCategory'] == status) for status in status_categories}
                
                # Calculate and store story points for each period and team member
                prev_story_points = sum(t['StoryPoints'] for t in prev_tasks if 'StoryPoints' in t)
                pre_prev_story_points = sum(t['StoryPoints'] for t in pre_prev_tasks if 'StoryPoints' in t)

                # No longer need to assign tracked time here, it's handled at the team_data level
                
                category_data['prev'][team_member] = prev_counts
                category_data['pre_prev'][team_member] = pre_prev_counts
                category_data['tasks']['prev'][team_member] = prev_tasks
                category_data['tasks']['pre_prev'][team_member] = pre_prev_tasks
                category_data['story_points']['prev'][team_member] = prev_story_points 
                category_data['story_points']['pre_prev'][team_member] = pre_prev_story_points 
                
            team_data[category] = category_data
        
        all_data[team_name] = team_data
    
    return all_data

def safe_set_cell_value(sheet, row, column, value):
    """Safely set cell value, handling merged cells properly"""
    # Get the cell
    cell = sheet.cell(row=row, column=column)
    
    # Check if it's a merged cell
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # Find the merge range that contains this cell
        for merge_range in sheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
            if min_row <= row <= max_row and min_col <= column <= max_col:
                # Set value in the top-left cell of the merged range
                sheet.cell(row=min_row, column=min_col, value=value)
                return
        
        # If we can't find the merge range, log a warning
        print(f"Warning: Cell at row {row}, column {column} is a merged cell but no merge range was found")
    else:
        # Set value directly for a normal cell
        cell.value = value


def create_consolidated_summary(wb, data):
    """Create a consolidated summary sheet that shows all teams together"""
    # This sheet is created by main() now, so just get it
    summary_sheet = wb["All Teams Summary"] 

    headers = ["Category", "Status"] + list(TEAMS.keys()) + ["Total"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    category_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    border_thick = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Find the starting row for the consolidated summary table.
    # It should be placed after the main summary content on the "Summary" sheet.
    # Let's assume it starts after the main summary, so we can find the max_row
    # of the "Summary" sheet and add some padding.
    summary_sheet_main = wb["Summary"]
    current_row_on_summary = summary_sheet_main.max_row + 3 # Add some space

    # Write a title for the consolidated summary table
    summary_sheet_main.cell(row=current_row_on_summary, column=1, value="Consolidated Status Table by Team").font = Font(bold=True, size=12)
    current_row_on_summary += 2 # Add more space before the table headers


    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header) # This is for "All Teams Summary" sheet
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    all_categories = sorted({cat for team_data in data.values() for cat in team_data.keys() if cat != 'aggregated_tracked_time'}) # Exclude the new key
    all_statuses = {"To Do", "In Development", "Completed", "Declined", "Cancelled"}

    row = 2
    for category in all_categories:
        summary_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        for col in range(1, len(headers) + 1):
            cell = summary_sheet.cell(row=row, column=col)
            cell.fill = category_fill
            cell.border = border_thick
            cell.alignment = left_align if col == 1 else center_align
        summary_sheet.cell(row=row, column=1).value = category
        summary_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for status in sorted(all_statuses):
            status_total = 0
            for col in range(1, len(headers) + 1):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align

            summary_sheet.cell(row=row, column=2, value=status)
            for team_idx, team in enumerate(TEAMS.keys()):
                team_total = 0
                if category in data.get(team, {}): # Check if category exists in team_data
                    for member in TEAMS[team]:
                        # Access prev counts from the original data structure
                        count = data[team][category]['prev'].get(member, {}).get(status, 0)
                        team_total += count
                col = team_idx + 3
                if team_total > 0:
                    summary_sheet.cell(row=row, column=col, value=team_total)
                status_total += team_total
            if status_total > 0:
                summary_sheet.cell(row=row, column=len(headers), value=status_total)
            row += 1

    # Add TOTAL row at the bottom
    total_row = summary_sheet.max_row + 1
    summary_sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    summary_sheet.cell(row=total_row, column=1).fill = total_fill
    summary_sheet.cell(row=total_row, column=1).border = thin_border
    summary_sheet.cell(row=total_row, column=1).alignment = center_align


def create_xlsx_report(data, wb): 
    """Create Excel report from the data"""
    try:
        # Always get the "Summary" sheet by name
        sheet = wb["Summary"]
        
        # Define cell styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11, color="444444")
        normal_font = Font(size=11)
        bold_font = Font(bold=True, size=11)
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        indent_align = Alignment(horizontal='left', vertical='center', indent=2)
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        category_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        status_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        alternating_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        totals_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        team_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        declined_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cancelled_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row = 1
        
        # Process each team separately
        for team_name, team_data in data.items():
            # Add team header
            team_cell = sheet.cell(row=row, column=1, value=team_name)
            team_cell.font = Font(bold=True, size=14)
            team_cell.fill = team_fill
            
            # Determine number of columns needed for this team
            team_members = TEAMS[team_name]
            member_count = len(team_members)
            total_cols = member_count + 2  # +1 for description column, +1 for totals column
            
            # Merge cells for team header
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
            
            # Style the team header
            for col in range(1, total_cols + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align
            
            row += 1
            
            # Set column widths
            sheet.column_dimensions['A'].width = 25
            for col_idx in range(member_count + 1):
                col_letter = get_column_letter(col_idx + 2)
                sheet.column_dimensions[col_letter].width = 15
            
            # Set headers for team members and total column
            sheet.cell(row=row, column=1, value="Task Category / Status")
            for idx, member in enumerate(team_members):
                sheet.cell(row=row, column=idx + 2, value=member)
            sheet.cell(row=row, column=total_cols, value="Status Totals")
            
            # Apply styling to headers
            for col in range(1, total_cols + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
            
            row += 1
            
            # Get the status categories for this team
            status_mapping = TEAM_STATUS_MAPPINGS.get(team_name, STATUS_MAPPING)
            status_categories = ['To Do', 'In Development', 'Completed']
            if 'DECLINED' in status_mapping:
                status_categories.append('Declined')
            if 'CANCELLED' in status_mapping:
                    status_categories.append('Cancelled')
            
            # Filter out 'aggregated_tracked_time' when iterating through categories for display
            display_categories = [cat for cat in team_data.keys() if cat != 'aggregated_tracked_time']

            # Fill in task categories and their counts
            for category_idx, category in enumerate(display_categories):
                category_data = team_data[category] # Get the actual category data
                # Add a row for the category name
                category_cell = sheet.cell(row=row, column=1, value=category)
                category_cell.font = bold_font
                category_cell.fill = category_fill
                    
                # Merge cells across all columns for category header
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
                    
                # Style category row
                for col in range(1, total_cols + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if col == 1:
                        cell.alignment = left_align
                        
                row += 1
                
                # Add a row for each status 
                for status_idx, status in enumerate(status_categories):
                    # Add the status name with indentation
                    status_cell = sheet.cell(row=row, column=1, value=f"{status}")
                    status_cell.alignment = indent_align
                    
                    # Apply special fill if this is a special status
                    if status == 'Declined' or status == 'Cancelled':
                        status_cell.fill = declined_fill if status == 'Declined' else cancelled_fill
                    
                    # Track status total for this row
                    status_prev_total = 0
                    
                    # Add data for each team member
                    for member_idx, team_member in enumerate(team_members):
                        prev_counts = category_data['prev'].get(team_member, {})
                        
                        # Get the count for this status, default to 0 if not found
                        prev_count = prev_counts.get(status, 0)
                        
                        # Add to totals
                        status_prev_total += prev_count
                        
                        # Add the count to the cell for this team member
                        cell = sheet.cell(row=row, column=member_idx + 2, value=f"{prev_count}")
                        
                        # Apply special fill if this is a special status
                        if status == 'Declined':
                            cell.fill = declined_fill
                        elif status == 'Cancelled':
                            cell.fill = cancelled_fill
                    
                    # Add status total to the totals column - without brackets
                    total_cell = sheet.cell(row=row, column=total_cols, value=f"{status_prev_total}")
                    total_cell.alignment = center_align
                    total_cell.font = subheader_font
                    
                    if status == 'Declined':
                        total_cell.fill = declined_fill
                    elif status == 'Cancelled':
                        total_cell.fill = cancelled_fill
                    else:
                        total_cell.fill = totals_fill
                    
                    # Style the row
                    for col in range(1, total_cols + 1):
                        cell = sheet.cell(row=row, column=col)
                        if col < total_cols:  # Skip the totals column as it's styled above
                            cell.font = normal_font
                            cell.border = thin_border
                            
                            if col > 1:
                                cell.alignment = center_align
                            
                            # Apply alternating colors
                            if status_idx % 2 == 1 and status not in ['Declined', 'Cancelled']:
                                cell.fill = alternating_fill
                        else:
                            # Make sure the total cell has a border
                            cell.border = thin_border
                    
                    row += 1
            
            # Add total rows
            total_row = row
            total_cell = sheet.cell(row=total_row, column=1, value="TOTAL")
            total_cell.font = bold_font
            total_cell.fill = header_fill
            total_cell.border = thin_border
            
            # Calculate totals for each team member across all categories
            grand_total_prev = 0
            
            for member_idx, team_member in enumerate(team_members):
                member_prev_total = 0
                
                for category_key_for_total, category_val_for_total in team_data.items():
                    if category_key_for_total != 'aggregated_tracked_time': # Exclude the new key
                        prev_counts = category_val_for_total['prev'].get(team_member, {})
                        for status in status_categories:
                            member_prev_total += prev_counts.get(status, 0)
                
                # Add to grand totals
                grand_total_prev += member_prev_total
                
                # Add totals for this team member (without brackets)
                sheet.cell(row=total_row, column=member_idx + 2, value=f"{member_prev_total}")
            
            # Add grand total to the totals column
            grand_total_cell = sheet.cell(row=total_row, column=total_cols, value=f"{grand_total_prev}")
            grand_total_cell.font = bold_font
            grand_total_cell.alignment = center_align
            grand_total_cell.fill = header_fill
            grand_total_cell.border = thin_border

            # Insert 'Story Points' and 'Tracked Time' rows after TOTAL
            # Calculate total story points for the team for the previous sprint
            team_story_points_total = 0
            team_tracked_time_total = 0 

            for offset, label in enumerate(["Story Points", "Tracked Time"]):
                metric_row = total_row + 1 + offset
                for col_idx in range(1, total_cols + 1):
                    cell = sheet.cell(row=metric_row, column=col_idx)
                    ref_cell = sheet.cell(row=total_row, column=col_idx)

                    # Copy style from TOTAL row
                    if col_idx == 1 or col_idx == total_cols:
                        # Label or Status Totals column: copy TOTAL row style
                        cell.fill = copy(ref_cell.fill)
                        cell.font = copy(ref_cell.font)
                        cell.alignment = copy(ref_cell.alignment)
                        cell.border = copy(ref_cell.border)
                    else:
                        # Employee columns: style like TOTAL employee cells
                        cell.fill = header_fill
                        cell.font = bold_font
                        cell.alignment = center_align
                        cell.border = thin_border

                    if col_idx == 1:
                        cell.value = label
                    elif label == "Story Points":
                        # Calculate story points for each member
                        if col_idx <= member_count + 1: 
                            member_name = team_members[col_idx - 2] 
                            member_story_points = 0
                            # Sum story points from all categories for the current member
                            for category_key_sp, category_val_sp in team_data.items():
                                if category_key_sp != 'aggregated_tracked_time': # Exclude the new key
                                    member_story_points += category_val_sp['story_points']['prev'].get(member_name, 0.0)
                            cell.value = member_story_points
                            team_story_points_total += member_story_points
                        elif col_idx == total_cols: # Total story points for the team
                            cell.value = team_story_points_total
                    elif label == "Tracked Time": 
                        if col_idx <= member_count + 1: 
                            member_name = team_members[col_idx - 2] 
                            # Retrieve tracked time directly from the team_data's aggregated_tracked_time
                            member_tracked_time = team_data['aggregated_tracked_time']['prev'].get(member_name, 0.0)
                            cell.value = f"{member_tracked_time:.2f}" # Format to 2 decimal places
                            team_tracked_time_total += member_tracked_time
                        elif col_idx == total_cols: # Total tracked time for the team
                            cell.value = f"{team_tracked_time_total:.2f}" # Format to 2 decimal places
            
            # Style total cells
            for col in range(2, total_cols):
                cell = sheet.cell(row=total_row, column=col)
                cell.font = bold_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
            
            # Add spacing between teams
            row = total_row + 4
        
        # Add date information at the end of the report
        row += 1
        sheet.cell(row=row, column=1, value=f"Previous Sprint: {PREV_SPRINT_START} to {PREV_SPRINT_END}")
        row += 1
        sheet.cell(row=row, column=1, value=f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Create consolidated summary across all teams
        create_consolidated_summary(wb, data) 
        add_consolidated_status_table(wb) 
        
        # Save the workbook
        wb.save(OUTPUT_PATH)
        print(f"Report saved to {OUTPUT_PATH}")
        
    except Exception as e:
        print(f"Error creating Excel report: {e}")
        import traceback
        traceback.print_exc()
        raise

def create_detailed_sheets(wb, data):
    """Create detailed sheets for tasks by team and category"""
    # Create one sheet for all task details
    details_sheet = wb.create_sheet("Task Details")
    
    # Setup headers for "Task Details" sheet
    headers_details = ["Team", "Category", "Key", "Summary", "Status", "Assignee", "Period", "Story Points", "Time Spent (Hours)"] 
    for col, header in enumerate(headers_details, 1):
        cell = details_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Set column widths for "Task Details" sheet
    details_sheet.column_dimensions['A'].width = 15  # Team
    details_sheet.column_dimensions['B'].width = 15  # Category
    details_sheet.column_dimensions['C'].width = 12  # Key
    details_sheet.column_dimensions['D'].width = 50  # Summary
    details_sheet.column_dimensions['E'].width = 15  # Status
    details_sheet.column_dimensions['F'].width = 15  # Assignee
    details_sheet.column_dimensions['G'].width = 15  # Period
    details_sheet.column_dimensions['H'].width = 15  # Story Points column
    details_sheet.column_dimensions['I'].width = 18 
    
    # Add data to "Task Details" sheet
    row_details = 2
    for team_name, team_data in data.items():
        # Filter out 'aggregated_tracked_time' when iterating through categories for display
        display_categories = [cat for cat in team_data.keys() if cat != 'aggregated_tracked_time']

        for category in display_categories:
            category_data = team_data[category] # Get the actual category data
            for team_member in TEAMS[team_name]:
                for task in category_data['tasks']['prev'].get(team_member, []):
                    details_sheet.cell(row=row_details, column=1, value=team_name)
                    details_sheet.cell(row=row_details, column=2, value=category)
                    details_sheet.cell(row=row_details, column=3, value=task['Key'])
                    details_sheet.cell(row=row_details, column=4, value=task['Summary'])
                    details_sheet.cell(row=row_details, column=5, value=task['Status'])
                    details_sheet.cell(row=row_details, column=6, value=task['Assignee'])
                    details_sheet.cell(row=row_details, column=7, value="Previous Sprint")
                    details_sheet.cell(row=row_details, column=8, value=task.get('StoryPoints', 0.0))
                    # Time spent per task within the period is not easily available here from the aggregated data
                    details_sheet.cell(row=row_details, column=9, value=f"{0.0:.2f}") 
                    row_details += 1
    
    # Create team-specific sheets
    for team_name, team_data in data.items():
        team_sheet = wb.create_sheet(f"{team_name}")
        
        # Setup headers for team-specific sheets (including hidden columns for data integrity)
        headers_team = ["Category", "Team Member", "Status", "Key", "Summary", "Story Points", "Time Spent (Hours)"] 
        for col, header in enumerate(headers_team, 1):
            cell = team_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Set column widths for team-specific sheets
        team_sheet.column_dimensions['A'].width = 15  # Category
        team_sheet.column_dimensions['B'].width = 15  # Team Member
        team_sheet.column_dimensions['C'].width = 15  # Status
        team_sheet.column_dimensions['D'].width = 12  # Key
        team_sheet.column_dimensions['E'].width = 50  # Summary
        team_sheet.column_dimensions['F'].width = 15  # Story Points column
        team_sheet.column_dimensions['G'].width = 18  
        
        # HIDE COLUMNS D, E, F, G on team-specific sheets
        team_sheet.column_dimensions['D'].hidden = True
        team_sheet.column_dimensions['E'].hidden = True
        team_sheet.column_dimensions['F'].hidden = True
        team_sheet.column_dimensions['G'].hidden = True 

        # Get the status categories for this team
        status_mapping = TEAM_STATUS_MAPPINGS.get(team_name, STATUS_MAPPING)
        status_categories = ['To Do', 'In Development', 'Completed']
        if 'DECLINED' in status_mapping:
            status_categories.append('Declined')
        if 'CANCELLED' in status_mapping:
            status_categories.append('Cancelled')
        
        # Add data
        row = 2
        
        # Filter out 'aggregated_tracked_time' when iterating through categories for display
        display_categories = [cat for cat in team_data.keys() if cat != 'aggregated_tracked_time']

        for category in display_categories:
            category_data = team_data[category] # Get the actual category data
            # Add category header
            cell = team_sheet.cell(row=row, column=1, value=category)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
            # Adjusted merge range to only cover visible columns (A, B, C)
            team_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3) 
            row += 1
            
            for team_member in TEAMS[team_name]:
                # Add team member header
                cell = team_sheet.cell(row=row, column=2, value=team_member)
                cell.font = Font(bold=True)
                row += 1
                
                # Add sections for each status category
                for status in status_categories:
                    cell = team_sheet.cell(row=row, column=3, value=status)
                    cell.font = Font(italic=True)
                    
                    # Apply special formatting for special statuses
                    if status == 'Declined':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif status == 'Cancelled':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    row += 1
                    
                    # Previous sprint tasks for this status - Data is still written here
                    # but columns D, E, F, G are hidden, so it's not visible to the user.
                    status_tasks = [t for t in category_data['tasks']['prev'].get(team_member, []) if t['StatusCategory'] == status]
                    if status_tasks:
                        for task in status_tasks:
                            team_sheet.cell(row=row, column=4, value=task['Key'])
                            team_sheet.cell(row=row, column=5, value=task['Summary'])
                            team_sheet.cell(row=row, column=6, value=task.get('StoryPoints', 0.0))
                            # Time spent per task within the period is not easily available here from the aggregated data
                            team_sheet.cell(row=row, column=7, value=f"{0.0:.2f}") 
                            row += 1
                    else:
                        # Still write "No tasks" to column 4 (hidden) to maintain structure
                        team_sheet.cell(row=row, column=4, value="No tasks")
                        row += 1
                
                row += 1  # Add space between team members
            
            row += 1  # Add space between categories

def main():
    print("\n--- Entering main function ---") # Added print statement here
    
    # Show team members
    for team_name, members in TEAMS.items():
        print(f"\n{team_name} Members:")
        for member in members:
            print(f"- {member}")
    
    # Connect to Jira
    jira = connect_to_jira()
    
    # Process all data
    data = process_data(jira)
    
    # Create a new workbook once at the start of main
    wb = openpyxl.Workbook()
    
    # Remove the default sheet created by openpyxl.Workbook() to avoid "Sheet"
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Explicitly create the "Summary" sheet first
    summary_sheet = wb.create_sheet("Summary", 0) 

    # Create the "All Teams Summary" sheet
    all_teams_summary_sheet = wb.create_sheet("All Teams Summary")

    # Call create_detailed_sheets BEFORE create_xlsx_report
    create_detailed_sheets(wb, data)
    
    # Pass the workbook to create_xlsx_report
    create_xlsx_report(data, wb) 
    
    print("Done!")


def add_consolidated_status_table(wb):
    """Add a consolidated status table to the 'Summary' sheet, using same styling."""
    summary_sheet = wb["Summary"]

    # Find last row and insert title
    last_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=last_row, column=1, value="Consolidated Status Table by Team").font = Font(bold=True, size=12)
    last_row += 2

    # Collect data
    status_data = []
    # Iterate through all team-specific task detail sheets
    for sheet_name in wb.sheetnames:
        # Only process sheets that are team-specific, not the summary or details sheets
        if sheet_name in ["Summary", "Task Details", "All Teams Summary"]:
            continue
        ws = wb[sheet_name]
        rows = list(ws.values)
        if not rows:
            continue
        # Headers are dynamic, look for 'Key' and 'Status'
        # Assuming the first row of team-specific sheets contains headers
        headers = [cell.value for cell in ws[1]] 
        
        # Find the column indices for 'Key' and 'Status'
        try:
            key_col_idx = headers.index("Key")
            status_col_idx = headers.index("Status")
        except ValueError:
            print(f"Skipping sheet '{sheet_name}' as 'Key' or 'Status' column not found.")
            continue

        # Extract data, skipping header row (row 1) and empty rows
        sheet_rows_data = []
        for r_idx in range(2, ws.max_row + 1): # Start from row 2 (after header)
            row_data = [cell.value for cell in ws[r_idx]]
            # Ensure row_data has enough elements before accessing indices
            if len(row_data) > max(key_col_idx, status_col_idx) and row_data[key_col_idx] is not None:
                sheet_rows_data.append({
                    'Key': row_data[key_col_idx],
                    'Status': row_data[status_col_idx]
                })

        if sheet_rows_data:
            df = pd.DataFrame(sheet_rows_data)
            df["Team"] = sheet_name.replace(" TEAM", "") # Clean up team name for pivot
            status_data.append(df)

    if not status_data:
        print("No data found for consolidated table.")
        return

    combined_df = pd.concat(status_data)
    pivot = (
        combined_df.groupby(["Team", "Status"])
        .size()
        .reset_index(name="Count")
        .pivot(index="Team", columns="Status", values="Count")
        .fillna(0)
        .astype(int)
        .reset_index()
    )

    # Style setup
    header_font = Font(bold=True, size=11)
    cell_font = Font(size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Write headers
    for col_idx, col_name in enumerate(pivot.columns, 1):
        cell = summary_sheet.cell(row=last_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Write data
    for r_idx, row in enumerate(pivot.itertuples(index=False), start=last_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

    # Add total row
    total_row = summary_sheet.max_row + 1
    summary_sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    summary_sheet.cell(row=total_row, column=1).fill = total_fill

    for col in range(2, pivot.shape[1] + 1):
        total = 0
        for r in range(last_row + 1, total_row):
            val = summary_sheet.cell(row=r, column=col).value
            if isinstance(val, int):
                total += val
        cell = summary_sheet.cell(row=total_row, column=col, value=total)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border

    print("✅ Consolidated styled table added to Summary.")


if __name__ == "__main__":
    main()
